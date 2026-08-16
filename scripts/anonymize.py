# -*- coding: utf-8 -*-
"""
sales-viz-secure · 数据脱敏与标准化（多行业通用版）

把任意销售/收入明细（xlsx / csv / json）转成看板可用的标准 JSON，同时按等级脱敏。
列名自动识别，支持不同行业的表头；通过 --industry 切换行业预置，默认 generic（安全中性）。

用法:
  python anonymize.py -i 明细.xlsx -o data_masked.json --level L2            # 通用行业（默认）
  python anonymize.py -i 明细.xlsx -o data_masked.json --industry education # 你们的培训行业
  python anonymize.py -i a.xlsx b.xlsx -o out.json --level L3 --scale 0.9
  python anonymize.py -i in.json -o out.json --level L2 --mapping mapping.json

行业预置（references/industry/*.json，可用 --industry 指定，默认 generic）：
  generic   中性标签、不碰科目/支付，任意行业开箱即用（安全默认）
  education K12→兴趣课、支付归四类、老师化名（你们现用的这套）
  retail    餐饮零售：商品原名、支付含 其他 兜底
  clinic    医疗诊所：患者编号、医师化名、支付含医保
  b2b       B2B制造：产品原名、支付含 月结/对公/信用证 等
  （也可传任意 json 文件路径作为自定义预置）

脱敏等级:
  L0  原样输出（仅内部核对用，不要外发）
  L1  客户姓名掩码（李婉贞 → 李*贞）；员工化名；科目按预置处理
  L2  客户编号化（→ 客户001 / 学员001）；员工化名；签单人留真名  ← 默认，给老板/管理层
  L3  客户编号化 + 员工代号化 + 可选金额缩放 + 可选日期偏移（对外/案例分享）

映射表:
  --mapping-out mapping.json  导出「编号/化名/科目 ↔ 真值」对照表（本地留存，切勿随看板外发）
  --mapping     mapping.json  读入已有对照表，保证多期数据编号一致
"""
import argparse
import csv
import json
import os
import re
import sys
from collections import OrderedDict
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from errors import EXIT_OK, EXIT_USAGE, EXIT_CONFIG, EXIT_RUNTIME, err, warn, info  # noqa: E402

# ---------------- 列名自动识别 ----------------
# role: date/amount/qty/customer/staff/dim
COLUMN_HINTS = [
    ('d',       'date',     ['日期', '报名日期', '签单日期', '缴费日期', '收款日期', '成交日期', '时间', 'date']),
    ('a',       'amount',   ['金额', '实收', '实收金额', '收入', '缴费金额', '价税合计', '成交金额', '费用', 'amount']),
    ('se',      'qty',      ['课时', '课次', '节数', '数量', '课时数', 'qty', 'sessions']),
    ('n',       'customer', ['姓名', '学员', '学员姓名', '客户', '客户名称', '客户姓名', '患者', '病人', '会员', '顾客', '会员名', 'name']),
    ('signer',  'staff',    ['签单人', '销售', '业务员', '顾问', '销售顾问', '成交人', '负责人', '跟单', '跟单员']),
    ('teacher', 'staff',    ['上课老师', '老师', '授课老师', '试听老师', '任课老师', '服务人员',
                             '店员', '技师', '医师', '护士', '导购', '客户经理', '服务员']),
    ('c',       'dim',      ['报名科目', '科目', '课程', '项目', '产品', '商品', '品类', '病种', '诊断']),
    ('src',     'dim',      ['来源', '续费新签', '客户来源', '渠道', '生源', '新老生', '推广']),
    ('p',       'dim',      ['支付方式', '付款方式', '收款方式', '支付渠道', '收款账户']),
    ('campus',  'dim',      ['校区', '所属校区', '门店', '分店', '部门', '院区', '分院']),
    # 地区层级（用于看板省→市→县→区 钻取；值非 PII，原样保留）
    ('prov',    'region',   ['省份', '省', '直辖市', '自治区', 'province']),
    ('city',    'region',   ['城市', '市', '地市', 'city']),
    ('county',  'region',   ['县', '县级', 'county']),
    ('dist',    'region',   ['区', '区县', '城区', '行政区', '街道', '乡镇', 'district']),
    ('region',  'region',   ['地区', '区域', '省市区', '省市区县', '归属地']),
    # 国家（用于世界地图；值非 PII，原样保留，中文名自动映射为英文）
    ('country', 'region',   ['国家', '国别', 'country', 'nation']),
    # 人物画像
    ('gender',  'demographic', ['性别', '男女', 'gender', 'sex']),
    ('age',     'demographic', ['年龄', '岁', '年纪', '龄', 'age']),
    # 漏斗/流程阶段（用于看板转化漏斗；值为阶段名，原样保留，不进维度图表）
    ('stage',   'dim', ['阶段', '状态', '环节', '流程阶段', '销售阶段', '漏斗阶段', 'stage', 'status', 'funnel_stage']),
    # 桑基图源/目标节点（用于看板流向桑基图；ssrc=上游/来源节点，star=下游/去向节点；
    #   维度遍历列表不含这两 key，故不会进维度排行，仅服务桑基图；值为节点名，原样保留）
    ('ssrc',    'dim', ['源节点', '上游节点', '父节点', '起始节点', '桑基源', '来源节点', 'source_node', 'from_node']),
    ('star',    'dim', ['目标节点', '下游节点', '子节点', '终点节点', '去向节点', '桑基目标', 'target_node', 'to_node']),
]
ROLE_OF = {k: r for k, r, _ in COLUMN_HINTS}
# 默认维度标签（generic 预置会覆盖）
DEFAULT_LABELS = {
    'd': '日期', 'a': '金额', 'se': '数量', 'n': '客户',
    'signer': '签单人', 'teacher': '服务人员', 'c': '项目',
    'src': '来源', 'p': '支付方式', 'campus': '门店/部门',
    'country': '国家',
}

# 行业预置目录
INDUSTRY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            '..', 'references', 'industry')


def load_industry(name):
    """载入行业预置 JSON；name=generic/education/retail/clinic/b2b，或任意文件路径。找不到回退 generic。"""
    if not name or name == 'generic':
        path = os.path.join(INDUSTRY_DIR, 'generic.json')
    else:
        cand = [os.path.join(INDUSTRY_DIR, name + '.json'),
                name if os.path.exists(name) else None,
                os.path.join(INDUSTRY_DIR, name)]
        path = next((c for c in cand if c and os.path.exists(c)), None)
    if not path or not os.path.exists(path):
        print('  ⚠ 未找到行业预置 %r，回退到 generic' % name)
        path = os.path.join(INDUSTRY_DIR, 'generic.json')
    with open(path, encoding='utf-8') as fp:
        cfg = json.load(fp)
    # 解析科目重映射文件（相对 references/industry/）
    if cfg.get('subject_remap_file'):
        sp = os.path.join(INDUSTRY_DIR, cfg['subject_remap_file'])
        cfg['_subject_remap'] = json.load(open(sp, encoding='utf-8')) if os.path.exists(sp) else None
    else:
        cfg['_subject_remap'] = None
    return cfg


def norm_payment(v, remap=None, keywords=None, unknown='__keep__'):
    """把任意支付方式写法归并；remap=None 时原样保留（不归并）。unknown='__keep__' 保留原值，否则归入该档。"""
    s = str(v).strip()
    if not s or s == '未提供':
        return '未提供'
    if remap:
        if s in remap:
            return remap[s]
        if keywords:
            for kw, target in keywords:
                if kw in s:
                    return target
        return unknown if unknown != '__keep__' else s
    return s


def guess_key(header: str):
    """把原始列名映射到标准字段 key"""
    h = str(header).strip()
    if not h:
        return None
    if h in ROLE_OF:          # 已经是标准 key，直接透传
        return h
    hl = h.lower()
    for key, _role, words in COLUMN_HINTS:      # 1) 完全相等
        for w in words:
            if hl == w.lower():
                return key
    for key, _role, words in COLUMN_HINTS:      # 2) 关键词出现在列名中（关键词需≥2字符，避免 a/d 误配）
        for w in words:
            if len(w) >= 2 and w.lower() in hl:
                return key
    return None


# ---------------- 读取 ----------------
def read_xlsx(path, sheet=None):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # 找表头行：非空单元格最多的前 5 行之一
    head_idx = max(range(min(5, len(rows))), key=lambda i: sum(1 for c in rows[i] if c not in (None, '')))
    header = [str(c).strip() if c is not None else '' for c in rows[head_idx]]
    out = []
    for r in rows[head_idx + 1:]:
        if all(c in (None, '') for c in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r))) if header[i]})
    return out


def read_csv(path):
    for enc in ('utf-8-sig', 'gbk', 'utf-8'):
        try:
            with open(path, encoding=enc, newline='') as fp:
                return list(csv.DictReader(fp))
        except UnicodeDecodeError:
            continue
    raise RuntimeError('无法识别 CSV 编码: ' + path)


def read_json(path):
    with open(path, encoding='utf-8') as fp:
        d = json.load(fp)
    if isinstance(d, dict):
        for v in d.values():
            if isinstance(v, list):
                return v
        raise RuntimeError('JSON 中未找到记录数组')
    return d


def read_sqlite(path, table=None):
    """读取 SQLite 库：--table 指定表名，否则取第一张业务表。"""
    import sqlite3
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        tables = [r[0] for r in cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'").fetchall()]
        if not tables:
            raise RuntimeError('SQLite 库中无业务表')
        t = table or tables[0]
        if table and table not in tables:
            raise RuntimeError('SQLite 库中无表 %r（现有：%s）' % (table, '、'.join(tables)))
        return [dict(r) for r in cur.execute('SELECT * FROM "%s"' % t).fetchall()]
    finally:
        conn.close()


def read_url(url):
    """http(s) URL 拉取 JSON/CSV（离线场景自动报错提示）。"""
    import urllib.request
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=30) as r:
        text = r.read().decode('utf-8')
    try:
        d = json.loads(text)
    except Exception:
        import io
        return list(csv.DictReader(io.StringIO(text)))
    if isinstance(d, dict):
        for v in d.values():
            if isinstance(v, list):
                return v
        raise RuntimeError('URL JSON 中未找到记录数组')
    return d


def read_any(path, sheet=None):
    if path == '-':
        # stdin 管道：JSON 或 CSV 自动识别（如 cat data.json | anonymize -i -）
        import io
        text = sys.stdin.read()
        try:
            d = json.loads(text)
        except Exception:
            return list(csv.DictReader(io.StringIO(text)))
        if isinstance(d, dict):
            for v in d.values():
                if isinstance(v, list):
                    return v
            raise RuntimeError('stdin JSON 中未找到记录数组')
        return d
    if path.startswith('http://') or path.startswith('https://'):
        # URL 拉取（P2 多源接入）
        return read_url(path)
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xlsm', '.xls'):
        return read_xlsx(path, sheet)
    if ext == '.csv':
        return read_csv(path)
    if ext == '.json':
        return read_json(path)
    if ext in ('.db', '.sqlite', '.sqlite3'):
        return read_sqlite(path, sheet)
    raise RuntimeError('不支持的文件类型: ' + ext + '（支持 xlsx/csv/json/sqlite/URL/stdin）')


# ---------------- 值清洗 ----------------
def norm_date(v, shift_days=0):
    if v is None or v == '':
        return ''
    if isinstance(v, datetime):
        dt = v
    else:
        s = str(v).strip().split(' ')[0]
        s = s.replace('/', '-').replace('.', '-').replace('年', '-').replace('月', '-').replace('日', '')
        m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
        if m:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        else:
            m2 = re.match(r'^(\d{1,2})-(\d{1,2})$', s)
            if m2:
                dt = datetime(datetime.now().year, int(m2.group(1)), int(m2.group(2)))
            else:
                return str(v)
    if shift_days:
        dt = dt + timedelta(days=shift_days)
    return dt.strftime('%Y-%m-%d')


def norm_num(v):
    if v is None or v == '':
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[^\d.\-]', '', str(v))
    try:
        return float(s) if s not in ('', '-', '.') else 0
    except ValueError:
        return 0


def mask_name(name: str) -> str:
    s = str(name).strip()
    if len(s) <= 1:
        return s + '*'
    if len(s) == 2:
        return s[0] + '*'
    return s[0] + '*' * (len(s) - 2) + s[-1]


# ---------------- 主流程 ----------------
def build(records, level='L2', scale=1.0, date_shift=0, mapping=None,
          keep_staff=None, extra_map=None, anonymize_teacher=True,
          subject_remap=None, payment_remap=None, payment_keywords=None, payment_unknown='__keep__',
          label_of=None, customer_prefix='客户', staff_friendly_pool=None,
          staff_friendly_prefix='服务人员', staff_code=None, hide_signer=False):
    """records: list[dict] 原始行 → (标准记录列表, 映射表, 维度列表)"""
    mapping = mapping or {}
    mapping.setdefault('customer', {})
    mapping.setdefault('staff', {})
    mapping.setdefault('teacher_friendly', {})
    mapping.setdefault('subject', {})
    mapping.setdefault('payment', {})

    if keep_staff is None:
        keep_staff = (level != 'L3')

    # 1. 列名映射
    all_headers = OrderedDict()
    for r in records:
        for h in r.keys():
            all_headers.setdefault(h, 0)
    colmap = {}
    used = set()
    ignored = []
    for h in all_headers:
        k = (extra_map or {}).get(h) or guess_key(h)
        if k and k not in used:
            colmap[h] = k
            used.add(k)
        else:
            ignored.append(h)
    if ignored:
        print('  · 忽略列: ' + ', '.join(ignored) + '（如需保留请用 --map "列名=标准key"）')
    if 'd' not in used or 'a' not in used:
        raise RuntimeError('未识别到「日期」或「金额」列，请用 --map 手动指定，例如 --map "成交日=d" --map "实收款=a"\n'
                           '识别到的列: ' + ', '.join(f'{h}->{colmap.get(h, "忽略")}' for h in all_headers))

    # 2. 逐行转换
    out = []
    for r in records:
        rec = {}
        for h, k in colmap.items():
            v = r.get(h)
            role = ROLE_OF.get(k, 'dim')
            if role == 'date':
                rec[k] = norm_date(v, date_shift)
            elif role == 'amount':
                rec[k] = round(norm_num(v) * scale, 2)
            elif role == 'qty':
                rec[k] = int(norm_num(v))
            elif role == 'demographic' and k == 'age':
                rec[k] = norm_num(v)
            else:
                rec[k] = str(v).strip() if v not in (None, '') else '未提供'
        if not rec.get('d') or not rec.get('a'):
            continue
        out.append(rec)

    # 3. 脱敏
    def next_code(pool, prefix, width=3):
        return '%s%0*d' % (prefix, width, len(pool) + 1)

    staff_code = staff_code or {'teacher': '员工', 'signer': '员工'}
    for rec in out:
        # 客户姓名：L1 掩码 / L2+ 编号化
        if 'n' in rec and level != 'L0':
            raw = rec['n']
            if level == 'L1':
                rec['n'] = mask_name(raw)
            else:
                if raw not in mapping['customer']:
                    mapping['customer'][raw] = next_code(mapping['customer'], customer_prefix)
                rec['n'] = mapping['customer'][raw]
        # 科目：按行业预置重映射（generic 无 remap 则原样保留）
        if 'c' in rec and subject_remap and level != 'L0':
            raw = rec['c']
            new = subject_remap.get(raw, subject_remap.get('__default__'))
            if new is None:
                print('  · 科目未命中重映射，保持原值: ' + raw)
                new = raw
            mapping['subject'].setdefault(raw, new)
            rec['c'] = new
        # 授课/服务老师：L3 代号化；L1/L2 友好化名（有池用池，否则前缀+字母）；--keep-teacher/--keep-staff 保留真名
        if 'teacher' in rec and level != 'L0':
            raw = rec['teacher']
            if (level == 'L3') and not keep_staff:
                mk = 'teacher::' + raw
                if mk not in mapping['staff']:
                    same = [x for x in mapping['staff'] if x.startswith('teacher::')]
                    mapping['staff'][mk] = '%s%d' % (staff_code.get('teacher', '员工'), len(same) + 1)
                rec['teacher'] = mapping['staff'][mk]
            elif anonymize_teacher:
                if raw not in mapping['teacher_friendly']:
                    if staff_friendly_pool:
                        mapping['teacher_friendly'][raw] = staff_friendly_pool[len(mapping['teacher_friendly']) % len(staff_friendly_pool)]
                    else:
                        mapping['teacher_friendly'][raw] = '%s%s' % (staff_friendly_prefix, chr(ord('A') + len(mapping['teacher_friendly'])))
                rec['teacher'] = mapping['teacher_friendly'][raw]
            # 否则保持真名
        # 签单人：L3 或 --hide-signer 时代号化；否则保留真名供管理层看业绩
        if 'signer' in rec and level != 'L0' and (level == 'L3' or hide_signer):
            raw = rec['signer']
            mk = 'signer::' + raw
            if mk not in mapping['staff']:
                same = [x for x in mapping['staff'] if x.startswith('signer::')]
                mapping['staff'][mk] = '%s%s' % (staff_code.get('signer', '员工'), chr(ord('A') + len(same)))
            rec['signer'] = mapping['staff'][mk]
        # 支付方式：按行业预置归并（generic 无 remap 则原样保留）
        if 'p' in rec and level != 'L0':
            raw = rec['p']
            new = norm_payment(raw, payment_remap, payment_keywords, payment_unknown)
            mapping['payment'].setdefault(raw, new)
            rec['p'] = new

    # 4. 维度清单（用于看板自动生成图表）
    dims = []
    for k in ('c', 'src', 'signer', 'teacher', 'p', 'campus'):
        if any(k in r for r in out):
            dims.append({'key': k, 'label': (label_of or DEFAULT_LABELS).get(k, k)})

    # 5. 元信息：地区层级 / 性别 / 年龄（供看板做钻取与人物画像）
    meta = {'regionKeys': [], 'genderKey': None, 'ageKey': None}
    if any('country' in r for r in out):
        # 含国家列 → 世界地图模式（可带省/州实现国家内下钻）
        sub = [k for k in ('prov', 'city', 'county', 'dist') if any(k in r for r in out)]
        meta['regionKeys'] = ['country'] + sub
    else:
        order = ['prov', 'city', 'county', 'dist']
        rk = [k for k in order if any(k in r for r in out)]
        if rk:
            meta['regionKeys'] = rk
        elif any('region' in r for r in out):
            meta['regionKeys'] = ['region']
    if any('gender' in r for r in out):
        meta['genderKey'] = 'gender'
    if any('age' in r for r in out):
        meta['ageKey'] = 'age'

    out.sort(key=lambda r: r['d'])
    return out, mapping, dims, meta


def main():
    ap = argparse.ArgumentParser(description='销售数据脱敏与标准化')
    ap.add_argument('-i', '--input', nargs='+', required=False, help='输入文件（可多个）xlsx/csv/json，或 - 用 stdin')
    ap.add_argument('-o', '--output', required=False, help='输出标准 JSON')
    ap.add_argument('--wizard', action='store_true', help='交互向导：逐项提问生成参数（仅交互终端可用）')
    ap.add_argument('--sample', type=int, default=0, help='随机采样 N 条后脱敏（演示用，固定种子）')
    ap.add_argument('--sheet', default=None, help='Excel 工作表名')
    ap.add_argument('--level', default='L2', choices=['L0', 'L1', 'L2', 'L3'], help='脱敏等级，默认 L2')
    ap.add_argument('--industry', default='generic', help='行业预置：generic(默认)/education/retail/clinic/b2b，或自定义 json 路径')
    ap.add_argument('--scale', type=float, default=1.0, help='金额缩放系数（L3 常用，如 0.9）')
    ap.add_argument('--date-shift', type=int, default=0, help='日期整体平移天数')
    ap.add_argument('--keep-staff', action='store_true', help='强制保留员工真名')
    ap.add_argument('--hide-staff', action='store_true', help='强制员工代号化')
    ap.add_argument('--keep-teacher', action='store_true', help='保留授课老师真名（默认改为李老师/王老师…化名）')
    ap.add_argument('--hide-signer', action='store_true', help='签单人也代号化（签单人A/B…，彻底去除真名）')
    ap.add_argument('--subject-remap', default=None, help='科目→兴趣课重映射表 JSON（默认用内置 K12→兴趣 表）')
    ap.add_argument('--no-subject-remap', action='store_true', help='不做科目重映射，保留原始科目名')
    ap.add_argument('--no-payment-remap', action='store_true', help='不做支付方式归并，保留原始支付写法')
    ap.add_argument('--mapping', default=None, help='读入已有映射表（保证跨期编号一致）')
    ap.add_argument('--mapping-out', default=None, help='导出映射表路径')
    ap.add_argument('--map', action='append', default=[], help='手动列映射，格式 "原列名=标准key"')
    ap.add_argument('--dedup', action='store_true', help='按 客户+日期+科目+金额 去重')
    args = ap.parse_args()

    if args.wizard and sys.stdin.isatty():
        def _ask(prompt, default=None):
            if default:
                prompt = '%s [%s]' % (prompt, default)
            try:
                v = input(prompt + '：').strip()
            except (EOFError, KeyboardInterrupt):
                print('\n[错误] 向导被中断')
                raise SystemExit(EXIT_USAGE)
            return v or default
        if not args.input:
            inp = _ask('输入文件（多个用空格分隔，或 - 用 stdin）')
            args.input = inp.split()
        if not args.output:
            args.output = _ask('输出标准 JSON')
        ind = _ask('行业（generic/education/retail/clinic/b2b）', args.industry)
        args.industry = ind or args.industry
        lv = _ask('脱敏等级（L0/L1/L2/L3）', args.level)
        args.level = lv or args.level
        print()

    if not args.input:
        err('需要 -i 输入文件（或加 --wizard 交互向导）')
        return EXIT_USAGE
    if not args.output:
        err('需要 -o 输出文件（或加 --wizard 交互向导）')
        return EXIT_USAGE

    extra_map = {}
    for item in args.map:
        if '=' in item:
            a, b = item.split('=', 1)
            extra_map[a.strip()] = b.strip()

    records = []
    for path in args.input:
        try:
            rs = read_any(path, args.sheet)
        except Exception as e:
            err('读取失败 %s：%s' % (path, e))
            return EXIT_RUNTIME
        print(f'读取 {path if path == "-" else os.path.basename(path)}: {len(rs)} 行')
        records.extend(rs)

    if args.sample and 0 < args.sample < len(records):
        import random
        random.seed(42)
        src_len = len(records)
        records = random.sample(records, args.sample)
        info('大数据演示模式：随机采样 %d 条（源 %d 条，种子固定）' % (args.sample, src_len))

    mapping = None
    if args.mapping and os.path.exists(args.mapping):
        with open(args.mapping, encoding='utf-8') as fp:
            mapping = json.load(fp)
        print(f'复用映射表: 客户 {len(mapping.get("customer", {}))} 人')

    keep_staff = True if args.keep_staff else (False if args.hide_staff else None)
    anonymize_teacher = not args.keep_teacher
    hide_signer = args.hide_signer

    # 载入行业预置
    cfg = load_industry(args.industry)
    print('  · 行业预置: %s（%s）' % (args.industry, cfg.get('name', '')))
    label_of = {**DEFAULT_LABELS, **cfg.get('dim_labels', {})}
    customer_prefix = cfg.get('customer_prefix', '客户')
    staff_friendly_pool = cfg.get('staff_friendly_pool')
    staff_friendly_prefix = cfg.get('staff_friendly_prefix', cfg.get('staff_label', '服务人员'))
    staff_code = cfg.get('staff_code', {'teacher': '员工', 'signer': '员工'})
    # 科目重映射：--subject-remap 显式优先；否则用预置（--no-subject-remap 关闭）
    if args.subject_remap:
        subject_remap = json.load(open(args.subject_remap, encoding='utf-8'))
    elif args.no_subject_remap:
        subject_remap = None
    else:
        subject_remap = cfg.get('_subject_remap')
    # 支付归并：--no-payment-remap 关闭；否则用预置（无 remap 则原样保留）
    payment_remap = None if args.no_payment_remap else cfg.get('payment_remap')
    payment_keywords = None if args.no_payment_remap else cfg.get('payment_keywords')
    payment_unknown = cfg.get('payment_unknown', '__keep__')

    rows, mapping, dims, meta = build(records, args.level, args.scale, args.date_shift,
                                mapping, keep_staff, extra_map, anonymize_teacher,
                                subject_remap, payment_remap, payment_keywords, payment_unknown,
                                label_of, customer_prefix, staff_friendly_pool,
                                staff_friendly_prefix, staff_code, hide_signer)

    if args.dedup:
        seen, uniq = set(), []
        for r in rows:
            k = (r.get('n', ''), r.get('d', ''), r.get('c', ''), r.get('a', 0))
            if k in seen:
                continue
            seen.add(k)
            uniq.append(r)
        print(f'去重: {len(rows)} → {len(uniq)}')
        rows = uniq

    with open(args.output, 'w', encoding='utf-8') as fp:
        json.dump({'rows': rows, 'dims': dims, 'level': args.level, 'meta': meta}, fp, ensure_ascii=False, indent=1)

    total = sum(r['a'] for r in rows)
    print(f'✔ 输出 {args.output}: {len(rows)} 条 / ¥{total:,.0f} / 脱敏等级 {args.level}')
    print(f'  维度: {", ".join(d["label"] for d in dims)}')

    mo = args.mapping_out or args.mapping
    if mo and (mapping['customer'] or mapping['staff'] or mapping['teacher_friendly'] or mapping['subject'] or mapping['payment']):
        with open(mo, 'w', encoding='utf-8') as fp:
            json.dump(mapping, fp, ensure_ascii=False, indent=1)
        print(f'⚠ 映射表已存 {mo} —— 本地留存，切勿与看板一起外发')
        print(f'  含：客户 {len(mapping["customer"])} / 老师化名 {len(mapping["teacher_friendly"])} / 科目 {len(mapping["subject"])} / 员工代号 {len(mapping["staff"])} / 支付方式 {len(mapping["payment"])}')
    return EXIT_OK


if __name__ == '__main__':
    sys.exit(main())
